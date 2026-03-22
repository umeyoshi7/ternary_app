"""
タイムテーブル Excel 出力モジュール

出力シート:
  "タイムテーブル" : 工程リスト（開始時刻・終了時刻・所要時間）
  "Ganttチャート"  : 横軸=時間のガントチャート（1セル=30分）
"""

from __future__ import annotations

import io
import math
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from timetable.flow_reader import ManufacturingFlow, OPERATION_TYPES, resolve_schedule

# ── 操作タイプ別カラーパレット (RGB hex, no '#') ─────────────────────────
OP_COLORS: dict[str, str] = {
    "CHARGE":      "AED6F1",  # 水色
    "HEAT":        "F1948A",  # 赤系
    "COOL":        "85C1E9",  # 青系
    "REACTION":    "A9DFBF",  # 緑系
    "CONCENTRATE": "F9E79F",  # 黄系
    "FILTER":      "D7BDE2",  # 紫系
    "TRANSFER":    "FAD7A0",  # オレンジ系
    "WASH":        "A8D8EA",  # シアン系
    "OTHER":       "D5D8DC",  # グレー
}

HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT   = Font(bold=True, size=14)
BODY_FONT    = Font(size=10)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Gantt チャートの1セル幅（分）
GANTT_CELL_MIN = 30


def _minutes_to_hhmm(minutes: float) -> str:
    h = int(minutes) // 60
    m = int(minutes) % 60
    return f"{h:02d}:{m:02d}"


def _set_header(ws, row: int, col: int, value: str, width: float | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if width is not None:
        ws.column_dimensions[get_column_letter(col)].width = width


def _set_body(ws, row: int, col: int, value: Any, align=None, fill_color: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.alignment = align or LEFT
    cell.border = THIN_BORDER
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    return cell


# ---------------------------------------------------------------------------
# シート1: タイムテーブル（一覧表）
# ---------------------------------------------------------------------------

def _write_timetable_sheet(wb: openpyxl.Workbook, flow: ManufacturingFlow,
                            schedule: dict, start_hour: float = 8.0):
    ws = wb.active
    ws.title = "タイムテーブル"
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "バッチ製造タイムテーブル"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ヘッダー行
    headers = [
        ("工程番号", 8),
        ("工程名",   22),
        ("操作タイプ", 12),
        ("前工程",    10),
        ("時間決定",  10),
        ("開始時刻",  10),
        ("終了時刻",  10),
        ("所要時間(分)", 12),
        ("所要時間(h)", 12),
        ("備考",      24),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        _set_header(ws, 2, col_idx, h, w)
    ws.row_dimensions[2].height = 20

    # データ行
    start_offset = start_hour * 60  # 製造開始時刻（分）→実時刻変換用
    for r, step in enumerate(flow.steps, start=3):
        sch = schedule.get(step.step_no, {})
        start_min = sch.get("start", 0.0)
        end_min   = sch.get("end",   0.0)
        dur_min   = sch.get("duration", 0.0)

        start_real = _minutes_to_hhmm(start_offset + start_min)
        end_real   = _minutes_to_hhmm(start_offset + end_min)
        dur_h      = round(dur_min / 60, 2)

        fill = OP_COLORS.get(step.op_type, OP_COLORS["OTHER"])
        prev_str = ", ".join(str(p) for p in step.prev_steps) if step.prev_steps else "-"

        _set_body(ws, r, 1, step.step_no, CENTER, fill)
        _set_body(ws, r, 2, step.name,   LEFT,   fill)
        _set_body(ws, r, 3, step.op_label, CENTER, fill)
        _set_body(ws, r, 4, prev_str,    CENTER, fill)
        _set_body(ws, r, 5, step.time_method, CENTER, fill)
        _set_body(ws, r, 6, start_real,  CENTER, fill)
        _set_body(ws, r, 7, end_real,    CENTER, fill)
        _set_body(ws, r, 8, round(dur_min, 1), CENTER, fill)
        _set_body(ws, r, 9, dur_h,       CENTER, fill)
        _set_body(ws, r, 10, step.note,  LEFT,   fill)
        ws.row_dimensions[r].height = 18

    # 凡例
    legend_row = len(flow.steps) + 4
    ws.cell(row=legend_row, column=1, value="凡例").font = Font(bold=True)
    for i, (key, label) in enumerate(OPERATION_TYPES.items()):
        col = (i % 5) * 2 + 1
        row = legend_row + 1 + i // 5
        c = ws.cell(row=row, column=col, value=label)
        c.fill = PatternFill("solid", fgColor=OP_COLORS.get(key, "D5D8DC"))
        c.font = BODY_FONT
        c.border = THIN_BORDER
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, 12
        )

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# シート2: Ganttチャート
# ---------------------------------------------------------------------------

def _write_gantt_sheet(wb: openpyxl.Workbook, flow: ManufacturingFlow,
                       schedule: dict, start_hour: float = 8.0):
    ws = wb.create_sheet("Ganttチャート")
    ws.sheet_view.showGridLines = False

    if not schedule:
        return

    total_min = max(v["end"] for v in schedule.values())
    n_cells = math.ceil(total_min / GANTT_CELL_MIN) + 1

    # 固定列: 工程番号, 工程名, 操作タイプ, 所要時間(分)
    FIXED_COLS = 4
    COL_STEP   = 1
    COL_NAME   = 2
    COL_TYPE   = 3
    COL_DUR    = 4
    GANTT_START_COL = FIXED_COLS + 1

    # タイトル
    last_col = GANTT_START_COL + n_cells - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(row=1, column=1, value="Ganttチャート")
    t.font = TITLE_FONT
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # ヘッダー（固定列）
    for col, label, w in [
        (COL_STEP, "工程番号", 8),
        (COL_NAME, "工程名",  22),
        (COL_TYPE, "操作タイプ", 12),
        (COL_DUR,  "所要時間(分)", 12),
    ]:
        _set_header(ws, 2, col, label, w)

    # ヘッダー（時間軸）: 日付行（行2）と時刻行（行3）
    # 行2: 時間軸ラベル（XX:00 の位置のみ表示）
    current_day = 0
    for i in range(n_cells):
        col = GANTT_START_COL + i
        abs_min = start_hour * 60 + i * GANTT_CELL_MIN
        day = int(abs_min // (24 * 60))
        hour = int((abs_min % (24 * 60)) // 60)
        minute = int(abs_min % 60)

        ws.column_dimensions[get_column_letter(col)].width = 3.5

        # 行2: 時刻ラベル（00分のセルにのみ記載）
        if minute == 0:
            label = f"Day{day+1}\n{hour:02d}:00" if day != current_day else f"{hour:02d}:00"
            current_day = day
            c2 = ws.cell(row=2, column=col, value=label)
            c2.fill = HEADER_FILL
            c2.font = Font(color="FFFFFF", bold=True, size=8)
            c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c2.border = THIN_BORDER
        else:
            c2 = ws.cell(row=2, column=col)
            c2.fill = HEADER_FILL
            c2.border = THIN_BORDER

    ws.row_dimensions[2].height = 28

    # データ行
    for r, step in enumerate(flow.steps, start=3):
        sch = schedule.get(step.step_no, {})
        start_min = sch.get("start", 0.0)
        end_min   = sch.get("end",   0.0)
        dur_min   = sch.get("duration", 0.0)

        fill_color = OP_COLORS.get(step.op_type, OP_COLORS["OTHER"])

        _set_body(ws, r, COL_STEP, step.step_no,   CENTER)
        _set_body(ws, r, COL_NAME, step.name,       LEFT)
        _set_body(ws, r, COL_TYPE, step.op_label,   CENTER)
        _set_body(ws, r, COL_DUR,  round(dur_min, 1), CENTER)
        ws.row_dimensions[r].height = 20

        # Gantt バー描画
        start_cell = int(start_min / GANTT_CELL_MIN)
        end_cell   = math.ceil(end_min / GANTT_CELL_MIN)

        for ci in range(n_cells):
            col = GANTT_START_COL + ci
            cell = ws.cell(row=r, column=col)
            cell.border = Border(
                left=Side(style="hair"), right=Side(style="hair"),
                top=Side(style="hair"),  bottom=Side(style="hair"),
            )
            if start_cell <= ci < end_cell:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                # バーの中央セルに工程名を記載
                mid_ci = start_cell + (end_cell - start_cell) // 2
                if ci == mid_ci and (end_cell - start_cell) >= 2:
                    cell.value = step.name
                    cell.font  = Font(size=8, bold=True)
                    cell.alignment = CENTER
            else:
                cell.fill = PatternFill("solid", fgColor="F8F9FA")

    # 現在時刻ライン（オプション: 開始時刻の縦線を太く）
    for r in range(2, len(flow.steps) + 3):
        c = ws.cell(row=r, column=GANTT_START_COL)
        c.border = Border(
            left=Side(style="medium"), right=Side(style="hair"),
            top=Side(style="hair"),   bottom=Side(style="hair"),
        )

    ws.freeze_panes = ws.cell(row=3, column=GANTT_START_COL)


# ---------------------------------------------------------------------------
# 公開 API
# ---------------------------------------------------------------------------

def write_timetable_excel(
    flow: ManufacturingFlow,
    start_hour: float = 8.0,
) -> bytes:
    """
    ManufacturingFlow からタイムテーブル Excel を生成し bytes で返す。

    Parameters
    ----------
    flow       : ManufacturingFlow（duration_min が設定済みであること）
    start_hour : 製造開始時刻（時）。デフォルト 8:00

    Returns
    -------
    bytes (xlsx)
    """
    schedule = resolve_schedule(flow)

    wb = openpyxl.Workbook()
    _write_timetable_sheet(wb, flow, schedule, start_hour)
    _write_gantt_sheet(wb, flow, schedule, start_hour)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
