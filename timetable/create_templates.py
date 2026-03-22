"""
テンプレート Excel ファイル生成スクリプト
実行: python timetable/create_templates.py
"""

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path

TEMPLATE_DIR = Path(__file__).parent / "templates"
TEMPLATE_DIR.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT   = Font(color="666666", italic=True, size=9)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")


def _h(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _note(ws, row, col, val, color="FFF9C4"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = NOTE_FONT
    c.alignment = LEFT
    c.border = THIN


def create_flow_template():
    wb = openpyxl.Workbook()

    # ── シート1: フロー ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "フロー"
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "製造フロー定義シート"
    t.font = Font(bold=True, size=14)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # 説明行
    ws.merge_cells("A2:H2")
    d = ws["A2"]
    d.value = (
        "【記入方法】 工程番号は連番(1,2,3...)。前工程番号は複数の場合カンマ区切り(例: 1,2)。"
        "  時間決定: 「手動」or「計算」。手動の場合は手動時間(分)に数値を入力。"
    )
    d.font = NOTE_FONT
    d.fill = PatternFill("solid", fgColor="EBF5FB")
    d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ヘッダー
    headers = [
        ("工程番号", 8), ("工程名", 24), ("操作タイプ", 14),
        ("前工程番号", 12), ("時間決定", 10), ("手動時間(分)", 12),
        ("担当者", 12), ("備考", 28),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 3, col, h, w)
    ws.row_dimensions[3].height = 20

    # 操作タイプ選択肢（ドロップダウン用コメント）
    op_comment = ws.cell(row=4, column=3)
    op_comment.comment = None  # openpyxlのDataValidationを使用

    from openpyxl.worksheet.datavalidation import DataValidation
    dv_type = DataValidation(
        type="list",
        formula1='"CHARGE,HEAT,COOL,REACTION,CONCENTRATE,FILTER,TRANSFER,WASH,OTHER"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください",
    )
    dv_type.sqref = "C4:C100"
    ws.add_data_validation(dv_type)

    dv_time = DataValidation(
        type="list",
        formula1='"手動,計算"',
        showDropDown=False,
    )
    dv_time.sqref = "E4:E100"
    ws.add_data_validation(dv_time)

    # サンプルデータ
    sample_rows = [
        (1, "原料仕込み",    "CHARGE",      "",  "手動", 30,  "担当A", "溶剤・原料を仕込む"),
        (2, "加熱昇温",      "HEAT",        "1", "計算", "",  "担当A", "目標温度まで昇温"),
        (3, "反応",          "REACTION",    "2", "計算", "",  "担当A", "反応速度解析より推算"),
        (4, "冷却",          "COOL",        "3", "計算", "",  "担当B", "冷却水で冷却"),
        (5, "晶析",          "OTHER",       "4", "手動", 60,  "担当B", "自然冷却晶析"),
        (6, "ろ過",          "FILTER",      "5", "計算", "",  "担当B", "加圧ろ過"),
        (7, "洗浄",          "WASH",        "6", "手動", 30,  "担当B", "洗液3回"),
        (8, "濃縮",          "CONCENTRATE", "7", "計算", "",  "担当C", "溶媒回収"),
        (9, "移液・仕上げ",  "TRANSFER",    "8", "手動", 20,  "担当C", "製品タンクへ移液"),
    ]
    fill_colors = ["FDFEFE", "FEF9E7"]
    for r, row in enumerate(sample_rows, start=4):
        fill = PatternFill("solid", fgColor=fill_colors[r % 2])
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = CENTER if c != 2 else LEFT
            cell.border = THIN
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"

    # ── シート2: パラメータ ──────────────────────────────────────────────
    ws2 = wb.create_sheet("パラメータ")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value = "計算パラメータ入力シート"
    t2.font = Font(bold=True, size=14)
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:E2")
    d2 = ws2["A2"]
    d2.value = "【記入方法】 工程番号はフローシートと対応。パラメータ名・値・単位を記入。空欄はデフォルト値を使用。"
    d2.font = NOTE_FONT
    d2.fill = PatternFill("solid", fgColor="EBF5FB")
    d2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.row_dimensions[2].height = 24

    headers2 = [("工程番号", 10), ("工程名(参照用)", 20), ("パラメータ名", 24), ("値", 12), ("単位", 12)]
    for col, (h, w) in enumerate(headers2, 1):
        _h(ws2, 3, col, h, w)
    ws2.row_dimensions[3].height = 20

    # サンプルパラメータ（加熱工程・ろ過工程）
    param_rows = [
        # 加熱工程 (step 2)
        (2, "加熱昇温", "初期温度",        25,    "℃"),
        (2, "加熱昇温", "目標温度",        80,    "℃"),
        (2, "加熱昇温", "処理液質量",      500,   "kg"),
        (2, "加熱昇温", "比熱容量",        2.0,   "kJ/(kg·K)"),
        (2, "加熱昇温", "総括伝熱係数U",   300,   "W/(m²·K)"),
        (2, "加熱昇温", "伝熱面積A",       2.5,   "m²"),
        (2, "加熱昇温", "加熱媒体温度",    120,   "℃"),
        # 冷却工程 (step 4)
        (4, "冷却",     "初期温度",        80,    "℃"),
        (4, "冷却",     "目標温度",        20,    "℃"),
        (4, "冷却",     "処理液質量",      500,   "kg"),
        (4, "冷却",     "比熱容量",        2.0,   "kJ/(kg·K)"),
        (4, "冷却",     "総括伝熱係数U",   250,   "W/(m²·K)"),
        (4, "冷却",     "伝熱面積A",       2.5,   "m²"),
        (4, "冷却",     "冷却媒体温度",    10,    "℃"),
        # ろ過工程 (step 6)
        (6, "ろ過",     "ろ液量",          0.3,   "m³"),
        (6, "ろ過",     "ろ過面積A",       1.0,   "m²"),
        (6, "ろ過",     "圧差ΔP",          200,   "kPa"),
        (6, "ろ過",     "ケーク比抵抗α",   5e11,  "m/kg"),
        (6, "ろ過",     "ろ液粘度μ",       1e-3,  "Pa·s"),
        (6, "ろ過",     "スラリー固体濃度Cm", 50,  "kg/m³"),
        (6, "ろ過",     "媒体抵抗Rm",      1e10,  "1/m"),
        # 濃縮工程 (step 8)
        (8, "濃縮",     "蒸気供給量",      200,   "kg/h"),
        (8, "濃縮",     "圧力",            101.325, "kPa"),
        (8, "濃縮",     "目標留去率",      0.8,   "-"),
    ]
    fill_colors2 = ["FDFEFE", "FEF9E7"]
    for r, row in enumerate(param_rows, start=4):
        fill = PatternFill("solid", fgColor=fill_colors2[r % 2])
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = CENTER if c not in (2, 3) else LEFT
            cell.border = THIN
        ws2.row_dimensions[r].height = 18

    ws2.freeze_panes = "A4"

    # ── シート3: 操作タイプ一覧（参照用） ──────────────────────────────
    ws3 = wb.create_sheet("操作タイプ一覧")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    t3 = ws3["A1"]
    t3.value = "操作タイプ一覧（参照用）"
    t3.font = Font(bold=True, size=12)
    t3.alignment = CENTER
    ws3.row_dimensions[1].height = 24

    op_headers = [("タイプコード", 16), ("操作名（日本語）", 20), ("時間決定", 12), ("必要パラメータ", 48)]
    for col, (h, w) in enumerate(op_headers, 1):
        _h(ws3, 2, col, h, w)

    from timetable.flow_reader import OPERATION_TYPES
    op_params = {
        "CHARGE":      "手動",
        "HEAT":        "計算: 初期温度, 目標温度, 処理液質量, 比熱容量, U, A, 加熱媒体温度",
        "COOL":        "計算: 初期温度, 目標温度, 処理液質量, 比熱容量, U, A, 冷却媒体温度",
        "REACTION":    "計算: 反応速度定数k, 初期濃度, 目標転化率 / または 手動",
        "CONCENTRATE": "計算: 蒸気供給量(kg/h), 圧力(kPa), 目標留去率",
        "FILTER":      "計算: ろ液量, ろ過面積A, 圧差ΔP, ケーク比抵抗α, ろ液粘度μ, Cm, Rm",
        "TRANSFER":    "手動",
        "WASH":        "手動",
        "OTHER":       "手動 or 計算（備考に記載）",
    }
    for r, (key, label) in enumerate(OPERATION_TYPES.items(), start=3):
        from timetable.timetable_writer import OP_COLORS
        fill = PatternFill("solid", fgColor=OP_COLORS.get(key, "D5D8DC"))
        for c, val in enumerate([key, label,
                                  "手動" if op_params[key].startswith("手動") else "計算/手動",
                                  op_params[key]], 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = CENTER if c != 4 else LEFT
            cell.border = THIN
        ws3.row_dimensions[r].height = 20

    out_path = TEMPLATE_DIR / "flow_template.xlsx"
    wb.save(out_path)
    print(f"テンプレート生成完了: {out_path}")
    return out_path


if __name__ == "__main__":
    create_flow_template()
