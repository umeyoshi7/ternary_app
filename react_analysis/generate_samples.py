"""サンプルデータ生成スクリプト"""
import sys, math, os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, '.')

R = 8.314  # J/(mol·K)


def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def hfill():
    return PatternFill('solid', fgColor='4472C4')


def write_data_sheet(ws, headers, rows, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = hfill()
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[1].height = 32
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val if val is not None else '').border = thin_border()
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w


def add_meta_sheet(wb, name='テスト', substance='物質A', c0=1.0, temp=25.0, note='自動生成'):
    ws = wb.create_sheet('実験条件')
    rows = [
        ('ラベル', '値', '単位'),
        ('実験名', name, ''),
        ('反応物質', substance, ''),
        ('初期濃度', c0, 'mol/L'),
        ('反応温度', temp, '°C'),
        ('実験日', '2026-03-09', 'date'),
        ('担当者', 'テスト', ''),
        ('備考', note, ''),
    ]
    for ri, (l, v, u) in enumerate(rows, 1):
        for ci, val in enumerate([l, v, u], 1):
            c = ws.cell(ri, ci, val)
            c.border = thin_border()
            if ri == 1:
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = hfill()
                c.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10


HDRS_A = ['時間 (Time)\n(min)', '濃度_A (Concentration_A)\n(mol/L)',
          '温度 (Temperature)\n(°C)', '備考 (Notes)\n(-)']
HDRS_ABC = ['時間 (Time)\n(min)', '濃度_A (Concentration_A)\n(mol/L)',
            '濃度_B (Concentration_B)\n(mol/L)', '濃度_C (Concentration_C)\n(mol/L)',
            '温度 (Temperature)\n(°C)', '備考 (Notes)\n(-)']
WIDTHS_A   = [18, 22, 18, 12]
WIDTHS_ABC = [18, 22, 22, 22, 18, 12]

os.makedirs('sample_data', exist_ok=True)


# ==============================================================
# Sample 1: 単純1次反応 + 2%ノイズ (k=0.0234 min⁻¹)
# ==============================================================
np.random.seed(42)
times = [0, 5, 10, 15, 20, 30, 45, 60, 90, 120]
k_true = 0.0234
cA_clean = [math.exp(-k_true * t) for t in times]
cA_noisy = [max(c + np.random.normal(0, 0.02), 0.001) for c in cA_clean]

wb = openpyxl.Workbook()
ws = wb.active; ws.title = '実験データ'
rows = list(zip(times, [round(c, 5) for c in cA_noisy],
                [25.0] * len(times), [''] * len(times)))
write_data_sheet(ws, HDRS_A, rows, WIDTHS_A)
add_meta_sheet(wb, '単純1次反応_k=0.0234 (ノイズ2%)', '物質A', 1.0, 25.0, 'ノイズ2%付き合成データ')
wb.save('sample_data/sample1_simple_1st_order.xlsx')
print('Sample 1: 単純1次反応 (ノイズ付き) — 保存完了')

# ==============================================================
# Sample 2: 逐次反応 A→B→C (k1=0.05, k2=0.02 min⁻¹)
# ==============================================================
k1, k2 = 0.05, 0.02
times2 = [0, 5, 10, 15, 20, 30, 45, 60, 90, 120]

def seq_A(t): return math.exp(-k1 * t)
def seq_B(t): return k1 / (k2 - k1) * (math.exp(-k1 * t) - math.exp(-k2 * t))
def seq_C(t): return max(1 - seq_A(t) - seq_B(t), 0)

wb2 = openpyxl.Workbook()
ws2 = wb2.active; ws2.title = '実験データ'
rows2 = [(t, round(seq_A(t), 5), round(seq_B(t), 5), round(seq_C(t), 5), 25.0, '')
         for t in times2]
write_data_sheet(ws2, HDRS_ABC, rows2, WIDTHS_ABC)
add_meta_sheet(wb2, '逐次反応_k1=0.05_k2=0.02', 'A→B→C', 1.0, 25.0, '合成データ（ノイズなし）')
wb2.save('sample_data/sample2_sequential.xlsx')
print('Sample 2: 逐次反応 A→B→C (k1=0.05, k2=0.02) — 保存完了')

# ==============================================================
# Sample 3: 並列反応 A→B + A→C (k1=0.03, k2=0.01 min⁻¹)
# ==============================================================
k1p, k2p = 0.03, 0.01
k_total = k1p + k2p
times3 = [0, 5, 10, 15, 20, 30, 45, 60, 90, 120]

wb3 = openpyxl.Workbook()
ws3 = wb3.active; ws3.title = '実験データ'
rows3 = [(t,
          round(math.exp(-k_total * t), 5),
          round(k1p / k_total * (1 - math.exp(-k_total * t)), 5),
          round(k2p / k_total * (1 - math.exp(-k_total * t)), 5),
          25.0, '')
         for t in times3]
write_data_sheet(ws3, HDRS_ABC, rows3, WIDTHS_ABC)
add_meta_sheet(wb3, '並列反応_k1=0.03_k2=0.01', 'A→B+A→C', 1.0, 25.0, '合成データ')
wb3.save('sample_data/sample3_parallel.xlsx')
print('Sample 3: 並列反応 A→B + A→C (k1=0.03, k2=0.01) — 保存完了')

# ==============================================================
# Sample 4: 単純反応 複数温度 → Arrhenius (Ea=50kJ/mol, A=6e6)
# 各温度での k 値:
#   25°C: k≈0.0104 min⁻¹  35°C: k≈0.0201 min⁻¹
#   45°C: k≈0.0372 min⁻¹  55°C: k≈0.0664 min⁻¹
# → t=120min の最低濃度 ≈ 0.000348 (≠ 0) → 全温度でフィット可能
# ==============================================================
Ea, Af = 50000, 6e6
times4 = [0, 10, 20, 40, 60, 80, 100, 120]

wb4 = openpyxl.Workbook()
ws4 = wb4.active; ws4.title = '実験データ'
all_rows4 = []
for T_c in [25.0, 35.0, 45.0, 55.0]:
    T_K = T_c + 273.15
    kt = Af * math.exp(-Ea / (R * T_K))
    note_list = [f'{T_c}°C 開始'] + [''] * 6 + [f'{T_c}°C 終了']
    for t, nt in zip(times4, note_list):
        all_rows4.append((t, round(math.exp(-kt * t), 6), T_c, nt))
write_data_sheet(ws4, HDRS_A, all_rows4, WIDTHS_A)
add_meta_sheet(wb4, '複数温度_Arrhenius_Ea=50kJ', '物質A', 1.0, '25-55°C', 'Ea=50kJ/mol, A=6e6')
wb4.save('sample_data/sample4_multi_temp_arrhenius.xlsx')
print('Sample 4: 複数温度 Arrhenius (25/35/45/55°C, Ea=50kJ/mol) — 保存完了')

# ==============================================================
# Sample 5: 異なる時間点 (A/B/Cが別タイミングで測定)
# 逐次反応 k1=0.04, k2=0.015
# ==============================================================
k1s, k2s = 0.04, 0.015

def s_A(t): return math.exp(-k1s * t)
def s_B(t): return k1s / (k2s - k1s) * (math.exp(-k1s * t) - math.exp(-k2s * t))
def s_C(t): return max(1 - s_A(t) - s_B(t), 0)

times_A5 = [0, 10, 20, 40,     60,     90, 120]
times_B5 = [0,     20, 40, 50, 60,         120]
times_C5 = [0,         40,         80,     120]
all_times5 = sorted(set(times_A5) | set(times_B5) | set(times_C5))

wb5 = openpyxl.Workbook()
ws5 = wb5.active; ws5.title = '実験データ'
rows5 = []
for t in all_times5:
    cA_v = round(s_A(t), 5) if t in times_A5 else None
    cB_v = round(s_B(t), 5) if t in times_B5 else None
    cC_v = round(s_C(t), 5) if t in times_C5 else None
    rows5.append((t, cA_v, cB_v, cC_v, 25.0, ''))
write_data_sheet(ws5, HDRS_ABC, rows5, WIDTHS_ABC)
add_meta_sheet(wb5, '異時間点測定_逐次反応', 'A→B→C', 1.0, 25.0,
               f'A:{times_A5}\nB:{times_B5}\nC:{times_C5}')
wb5.save('sample_data/sample5_different_timepoints.xlsx')
print(f'Sample 5: 異時間点測定 ({len(rows5)}行, A:{len(times_A5)}点 B:{len(times_B5)}点 C:{len(times_C5)}点) — 保存完了')

# ==============================================================
# Sample 6: 複数温度 × 逐次反応 (k1/k2 両方でArrhenius)
# k1: Ea=50kJ/mol A=1e7  /  k2: Ea=30kJ/mol A=5e5
# ==============================================================
Ea1, Af1 = 50000, 1e7
Ea2, Af2 = 30000, 5e5
times6 = [0, 10, 20, 40, 60, 80, 100]

wb6 = openpyxl.Workbook()
ws6 = wb6.active; ws6.title = '実験データ'
rows6 = []
for T_c in [25.0, 40.0, 55.0]:
    T_K = T_c + 273.15
    k1t = Af1 * math.exp(-Ea1 / (R * T_K))
    k2t = Af2 * math.exp(-Ea2 / (R * T_K))
    for t in times6:
        cA_v = math.exp(-k1t * t)
        if abs(k2t - k1t) > 1e-10:
            cB_v = k1t / (k2t - k1t) * (math.exp(-k1t * t) - math.exp(-k2t * t))
        else:
            cB_v = k1t * t * math.exp(-k1t * t)
        cC_v = max(1 - cA_v - cB_v, 0)
        rows6.append((t, round(cA_v, 6), round(cB_v, 6), round(cC_v, 6), T_c, ''))
write_data_sheet(ws6, HDRS_ABC, rows6, WIDTHS_ABC)
add_meta_sheet(wb6, '複数温度×逐次反応', 'A→B→C', 1.0, '25/40/55°C',
               'k1: Ea=50kJ A=1e7  k2: Ea=30kJ A=5e5')
wb6.save('sample_data/sample6_multi_temp_sequential.xlsx')
print('Sample 6: 複数温度×逐次反応 (25/40/55°C, k1 Ea=50kJ, k2 Ea=30kJ) — 保存完了')

# ==============================================================
# Summary
# ==============================================================
print()
print('=' * 55)
print('生成完了: sample_data/ に6ファイル')
print('=' * 55)
for f in sorted(os.listdir('sample_data')):
    size = os.path.getsize(f'sample_data/{f}')
    print(f'  {f:45s} {size:6d} bytes')
