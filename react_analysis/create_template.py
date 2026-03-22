"""Script to create the experiment Excel and CSV templates."""

import math
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TEMPLATE_PATH     = os.path.join(os.path.dirname(__file__), "template", "experiment_template.xlsx")
CSV_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template", "experiment_template.csv")

# Sample data: sequential reaction A→B→C with k1=0.05, k2=0.02 min⁻¹ (approx)
SAMPLE_TIMES = [0, 5, 10, 15, 20, 30, 45, 60, 90, 120]
# A: C_A = exp(-k1*t), k1=0.05
SAMPLE_CONC_A = [round(math.exp(-0.05 * t), 4) for t in SAMPLE_TIMES]


# B: analytic for A→B→C
def _B(t, k1=0.05, k2=0.02):
    return k1 / (k2 - k1) * (math.exp(-k1 * t) - math.exp(-k2 * t))


SAMPLE_CONC_B = [round(_B(t), 4) for t in SAMPLE_TIMES]
# C: mass balance C = 1 - A - B
SAMPLE_CONC_C = [round(max(1.0 - SAMPLE_CONC_A[i] - SAMPLE_CONC_B[i], 0.0), 4) for i in range(len(SAMPLE_TIMES))]
SAMPLE_TEMP   = [25.0] * len(SAMPLE_TIMES)
SAMPLE_NOTES  = ["開始"] + [""] * (len(SAMPLE_TIMES) - 2) + ["終了"]


def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill():
    return PatternFill("solid", fgColor="4472C4")


def create_template():
    """Create Excel template with sample data."""
    wb = openpyxl.Workbook()

    # ---- Sheet 1: 実験データ ----
    ws1 = wb.active
    ws1.title = "実験データ"

    headers = [
        "時間 (Time)",
        "濃度_A (Concentration_A)",
        "濃度_B (Concentration_B)",
        "濃度_C (Concentration_C)",
        "温度 (Temperature)",
        "備考 (Notes)",
    ]
    units = ["min", "mol/L", "mol/L", "mol/L", "°C", "-"]

    # Header row
    for col, (h, u) in enumerate(zip(headers, units), start=1):
        cell = ws1.cell(row=1, column=col, value=f"{h}\n({u})")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill()
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = thin_border()

    ws1.row_dimensions[1].height = 32

    # Data rows
    for i, (t, cA, cB, cC, temp, note) in enumerate(
        zip(SAMPLE_TIMES, SAMPLE_CONC_A, SAMPLE_CONC_B, SAMPLE_CONC_C, SAMPLE_TEMP, SAMPLE_NOTES),
        start=2,
    ):
        ws1.cell(row=i, column=1, value=t).border = thin_border()
        ws1.cell(row=i, column=2, value=cA).border = thin_border()
        ws1.cell(row=i, column=3, value=cB).border = thin_border()
        ws1.cell(row=i, column=4, value=cC).border = thin_border()
        ws1.cell(row=i, column=5, value=temp).border = thin_border()
        ws1.cell(row=i, column=6, value=note).border = thin_border()

    # Column widths
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 24
    ws1.column_dimensions["D"].width = 24
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 16

    os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)
    wb.save(TEMPLATE_PATH)
    print(f"Excel template saved: {TEMPLATE_PATH}")


def create_csv_template():
    """Create CSV template with sample data and header."""
    lines = ["time,concentration,concentration_B,concentration_C,temperature,notes"]
    for i, (t, cA, cB, cC, temp, note) in enumerate(
        zip(SAMPLE_TIMES, SAMPLE_CONC_A, SAMPLE_CONC_B, SAMPLE_CONC_C, SAMPLE_TEMP, SAMPLE_NOTES)
    ):
        lines.append(f"{t},{cA},{cB},{cC},{temp},{note}")

    os.makedirs(os.path.dirname(CSV_TEMPLATE_PATH), exist_ok=True)
    with open(CSV_TEMPLATE_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")
    print(f"CSV template saved: {CSV_TEMPLATE_PATH}")


if __name__ == "__main__":
    create_template()
    create_csv_template()
